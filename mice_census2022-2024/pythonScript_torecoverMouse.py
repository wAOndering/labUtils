import pandas as pd
from openpyxl import load_workbook
import webbrowser


# Load the workbook and select the active worksheet
wb = load_workbook("C:/Users/Windows/Desktop/mice.xlsx")
ws = wb.active

# Initialize an empty list to store the data
data = []

# Loop through the cells in the worksheet
for row in ws.iter_rows():
    row_data = []
    for cell in row:
        # Check if the cell contains a hyperlink
        if cell.hyperlink:
            # Add the hyperlink to the row_data
            row_data.append(cell.hyperlink.target)
        else:
            # Add the regular cell value to the row_data
            row_data.append(cell.value)
    # Append the row_data to the data list
    data.append(row_data)

# Create a DataFrame from the data
df = pd.DataFrame(data)

# Print the DataFrame (optional) or work with it
print(df)


# Set the first row as the header and remove it from the DataFrame
df.columns = df.iloc[0]  # Set the first row as header
df = df.drop(0)  # Drop the first row

# Reset the index (optional, for cleaner output)
df = df.reset_index(drop=True)

# Split the column by space into two columns
df['invoice'] = df['invoice'].str.replace(u'\xa0', ' ') 
df[['Prefix', 'Number']] = df['invoice'].str.split(' ', expand=True)

df = df[df['Prefix']=='ACS']
dfsort = df.sort_values(by='date ')
dfsort = dfsort.reset_index(drop=True)

dfsort.to_csv('C:/Users/Windows/Desktop/mice_test.csv')
urls = dfsort.url.values
webbrowser.open(test)






from selenium import webdriver
from selenium.webdriver.chrome.options import Options



# Open a webpage
profile_path = r'C:\Users\Windows\AppData\Local\Google\Chrome\User Data\Profile 15'
chrome_options = Options()
chrome_options.add_argument("user-data-dir=C:/Users/Windows/AppData/Local/Google/Chrome/User Data/Profile 15")
driver = webdriver.Chrome(options=chrome_options)

for url in urls:
	driver.get(url)
	wait = WebDriverWait(driver, 5)

	# Find the element that triggers _doPostBack
	# You can use the ID or any other selector to find the button (adjust as per your page)
	link_button = wait.until(EC.element_to_be_clickable((By.ID, 'ctl00_ContentPlaceHolder1_LinkButton2')))

	# Click the button to trigger the JavaScript _doPostBack
	link_button.click()
